"""Generate a Turkish proforma invoice template (USD) for Genomer."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from copy import copy

OUT = "/Users/ilkerkadirozturk/Documents/genomer_brochures/genomer-proforma-taslak.xlsx"
LOGO = "/Users/ilkerkadirozturk/Documents/genomer_brochures/genomerlogo.png"

wb = Workbook()
ws = wb.active
ws.title = "Proforma Fatura"

# ---- Page setup ----
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5

# ---- Colors / fonts ----
BRAND = "0A5A8A"
LIGHT = "F5F8FB"
BORDER_COLOR = "B8C2CC"

thin = Side(border_style="thin", color=BORDER_COLOR)
medium = Side(border_style="medium", color=BRAND)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
heavy_bottom = Border(bottom=medium)

font_title = Font(name="Calibri", size=22, bold=True, color=BRAND)
font_subtitle = Font(name="Calibri", size=11, italic=True, color="555555")
font_company = Font(name="Calibri", size=11, bold=True, color=BRAND)
font_small = Font(name="Calibri", size=9, color="333333")
font_label = Font(name="Calibri", size=9, bold=True, color="555555")
font_value = Font(name="Calibri", size=10, color="111111")
font_section = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_th = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_td = Font(name="Calibri", size=10, color="111111")
font_total_label = Font(name="Calibri", size=10, bold=True, color="111111")
font_total_value = Font(name="Calibri", size=11, bold=True, color=BRAND)
font_footer = Font(name="Calibri", size=8, italic=True, color="666666")

fill_brand = PatternFill("solid", fgColor=BRAND)
fill_light = PatternFill("solid", fgColor=LIGHT)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---- Column widths (A..H) ----
widths = {"A": 6, "B": 32, "C": 14, "D": 8, "E": 10, "F": 14, "G": 14, "H": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---- Logo ----
try:
    img = Image(LOGO)
    img.height = 70
    img.width = 175
    ws.add_image(img, "A1")
except Exception as e:
    print(f"Logo skipped: {e}")

# Reserve rows for letterhead
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 20

# Company text on the right (rows 1-4, cols E-H)
ws.merge_cells("E1:H1")
ws["E1"] = "GENOMER BİYOTEKNOLOJİ ARGE SAN. VE TİC. LTD. ŞTİ."
ws["E1"].font = font_company
ws["E1"].alignment = right

ws.merge_cells("E2:H2")
ws["E2"] = "Barış SB Mah. 5003 Sk. Kadir Has Binası Kısım A No: 2, İç Kapı No: Z14"
ws["E2"].font = font_small
ws["E2"].alignment = right

ws.merge_cells("E3:H3")
ws["E3"] = "Gebze / Kocaeli / Türkiye"
ws["E3"].font = font_small
ws["E3"].alignment = right

ws.merge_cells("E4:H4")
ws["E4"] = "www.genomer.com.tr  •  support@genomer.com.tr"
ws["E4"].font = font_small
ws["E4"].alignment = right

# Letterhead bottom border on row 5
for col in "ABCDEFGH":
    ws[f"{col}5"].border = heavy_bottom
ws.row_dimensions[5].height = 4

# ---- Title row ----
ws.merge_cells("A6:H6")
ws["A6"] = "PROFORMA FATURA"
ws["A6"].font = font_title
ws["A6"].alignment = center
ws.row_dimensions[6].height = 36

ws.merge_cells("A7:H7")
ws["A7"] = "Proforma Invoice"
ws["A7"].font = font_subtitle
ws["A7"].alignment = center
ws.row_dimensions[7].height = 16

# ---- Meta block (proforma no / date / currency / validity) ----
meta_start = 9

def label_value(row, label_col, value_col_start, value_col_end, label, value=""):
    ws[f"{label_col}{row}"] = label
    ws[f"{label_col}{row}"].font = font_label
    ws[f"{label_col}{row}"].alignment = left
    if value_col_start == value_col_end:
        ws[f"{value_col_start}{row}"] = value
        ws[f"{value_col_start}{row}"].font = font_value
        ws[f"{value_col_start}{row}"].alignment = left
        ws[f"{value_col_start}{row}"].border = box
    else:
        ws.merge_cells(f"{value_col_start}{row}:{value_col_end}{row}")
        ws[f"{value_col_start}{row}"] = value
        ws[f"{value_col_start}{row}"].font = font_value
        ws[f"{value_col_start}{row}"].alignment = left
        for c in range(ord(value_col_start), ord(value_col_end) + 1):
            ws[f"{chr(c)}{row}"].border = box

label_value(meta_start,     "A", "B", "D", "Proforma No:")
label_value(meta_start,     "E", "F", "H", "Tarih:")
label_value(meta_start + 1, "A", "B", "D", "Para Birimi:", "USD")
label_value(meta_start + 1, "E", "F", "H", "Geçerlilik Tarihi:")

for r in range(meta_start, meta_start + 2):
    ws.row_dimensions[r].height = 22

# ---- Seller / Buyer blocks ----
sb_header = meta_start + 3  # row 12

ws.merge_cells(f"A{sb_header}:D{sb_header}")
ws[f"A{sb_header}"] = "SATICI / SELLER"
ws[f"A{sb_header}"].font = font_section
ws[f"A{sb_header}"].alignment = left
ws[f"A{sb_header}"].fill = fill_brand
for c in "ABCD":
    ws[f"{c}{sb_header}"].fill = fill_brand

ws.merge_cells(f"E{sb_header}:H{sb_header}")
ws[f"E{sb_header}"] = "ALICI / BUYER"
ws[f"E{sb_header}"].font = font_section
ws[f"E{sb_header}"].alignment = left
for c in "EFGH":
    ws[f"{c}{sb_header}"].fill = fill_brand

ws.row_dimensions[sb_header].height = 22

seller_lines = [
    "Genomer Biyoteknoloji Arge San. ve Tic. Ltd. Şti.",
    "Barış SB Mah. 5003 Sk. Kadir Has Binası",
    "Kısım A No: 2, İç Kapı No: Z14",
    "Gebze / Kocaeli / Türkiye",
    "Tel: +90 (___) ___ __ __",
    "E-posta: support@genomer.com.tr",
    "Vergi Dairesi / No: ____________________",
]
buyer_lines = [
    "Firma Adı / Company:",
    "Adres / Address:",
    "",
    "Ülke / Country:",
    "Tel:",
    "E-posta / Email:",
    "Vergi No / Tax ID:",
]

for i, (s, b) in enumerate(zip(seller_lines, buyer_lines)):
    r = sb_header + 1 + i
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = s
    ws[f"A{r}"].font = font_value
    ws[f"A{r}"].alignment = left_top
    ws.merge_cells(f"E{r}:H{r}")
    ws[f"E{r}"] = b
    ws[f"E{r}"].font = font_value
    ws[f"E{r}"].alignment = left_top
    ws.row_dimensions[r].height = 18
    for c in "ABCD":
        ws[f"{c}{r}"].fill = fill_light
    for c in "ABCDEFGH":
        ws[f"{c}{r}"].border = box

# ---- Items table ----
items_header_row = sb_header + 1 + len(seller_lines) + 1  # blank row, then header
ws.row_dimensions[items_header_row - 1].height = 8

headers = [
    ("A", "Sıra\nNo"),
    ("B", "Ürün Adı / Açıklama"),
    ("C", "GTIP / HS Code"),
    ("D", "Miktar"),
    ("E", "Birim"),
    ("F", "Birim Fiyat\n(USD)"),
    ("G", "Tutar\n(USD)"),
    ("H", "Açıklama"),
]
for col, label in headers:
    cell = ws[f"{col}{items_header_row}"]
    cell.value = label
    cell.font = font_th
    cell.fill = fill_brand
    cell.alignment = center
    cell.border = box
ws.row_dimensions[items_header_row].height = 32

ITEM_ROWS = 12
first_item = items_header_row + 1
last_item = items_header_row + ITEM_ROWS

for i in range(ITEM_ROWS):
    r = first_item + i
    ws[f"A{r}"] = i + 1
    ws[f"A{r}"].alignment = center
    ws[f"A{r}"].font = font_td
    for col in "BCDEFGH":
        ws[f"{col}{r}"].font = font_td
        ws[f"{col}{r}"].border = box
    ws[f"A{r}"].border = box
    ws[f"B{r}"].alignment = left
    ws[f"C{r}"].alignment = center
    ws[f"D{r}"].alignment = center
    ws[f"E{r}"].alignment = center
    ws[f"F{r}"].alignment = right
    ws[f"G{r}"].alignment = right
    ws[f"H{r}"].alignment = left
    ws[f"F{r}"].number_format = '#,##0.00 "USD"'
    ws[f"G{r}"].number_format = '#,##0.00 "USD"'
    ws[f"G{r}"] = f"=IF(AND(D{r}<>\"\",F{r}<>\"\"),D{r}*F{r},\"\")"
    if i % 2 == 1:
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].fill = fill_light
    ws.row_dimensions[r].height = 22

# ---- Totals ----
total_row = last_item + 1

def total_line(row, label, formula, bold=False):
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = font_total_label if not bold else font_total_value
    ws[f"A{row}"].alignment = right
    ws.merge_cells(f"G{row}:H{row}")
    ws[f"G{row}"] = formula
    ws[f"G{row}"].number_format = '#,##0.00 "USD"'
    ws[f"G{row}"].alignment = right
    ws[f"G{row}"].font = font_total_label if not bold else font_total_value
    for col in "ABCDEFGH":
        ws[f"{col}{row}"].border = box
    if bold:
        for col in "ABCDEFGH":
            ws[f"{col}{row}"].fill = fill_light
    ws.row_dimensions[row].height = 22

total_line(total_row,     "Ara Toplam / Subtotal:", f"=SUM(G{first_item}:G{last_item})")
total_line(total_row + 1, "İskonto / Discount:",     "")
total_line(total_row + 2, "Navlun / Freight:",        "")
total_line(total_row + 3, "GENEL TOPLAM / TOTAL:",   f"=G{total_row}-IFERROR(G{total_row+1},0)+IFERROR(G{total_row+2},0)", bold=True)

# ---- Terms block ----
terms_start = total_row + 5
ws.row_dimensions[terms_start - 1].height = 8

ws.merge_cells(f"A{terms_start}:H{terms_start}")
ws[f"A{terms_start}"] = "ŞARTLAR / TERMS & CONDITIONS"
ws[f"A{terms_start}"].font = font_section
ws[f"A{terms_start}"].alignment = left
for c in "ABCDEFGH":
    ws[f"{c}{terms_start}"].fill = fill_brand
ws.row_dimensions[terms_start].height = 20

terms = [
    ("Ödeme Şekli / Payment:", "Peşin (Cash in Advance)"),
    ("Teslim Şekli / Incoterms:", ""),
    ("Sevkiyat Süresi / Lead Time:", ""),
    ("Menşei / Country of Origin:", "Türkiye"),
    ("Ambalaj / Packaging:", ""),
    ("Notlar / Notes:", ""),
]
for i, (k, v) in enumerate(terms):
    r = terms_start + 1 + i
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"] = k
    ws[f"A{r}"].font = font_label
    ws[f"A{r}"].alignment = left
    ws.merge_cells(f"D{r}:H{r}")
    ws[f"D{r}"] = v
    ws[f"D{r}"].font = font_value
    ws[f"D{r}"].alignment = left
    for c in "ABCDEFGH":
        ws[f"{c}{r}"].border = box
    ws.row_dimensions[r].height = 20

# ---- Signature block ----
sig_row = terms_start + 1 + len(terms) + 2
ws.row_dimensions[sig_row - 1].height = 16

ws.merge_cells(f"A{sig_row}:D{sig_row}")
ws[f"A{sig_row}"] = "Düzenleyen / Issued By"
ws[f"A{sig_row}"].font = font_label
ws[f"A{sig_row}"].alignment = center

ws.merge_cells(f"E{sig_row}:H{sig_row}")
ws[f"E{sig_row}"] = "Yetkili İmza & Kaşe / Authorized Signature & Stamp"
ws[f"E{sig_row}"].font = font_label
ws[f"E{sig_row}"].alignment = center

ws.merge_cells(f"A{sig_row+1}:D{sig_row+1}")
ws[f"A{sig_row+1}"] = "Genomer Biyoteknoloji Arge San. ve Tic. Ltd. Şti."
ws[f"A{sig_row+1}"].font = font_value
ws[f"A{sig_row+1}"].alignment = center

ws.merge_cells(f"E{sig_row+1}:H{sig_row+1}")
ws[f"E{sig_row+1}"] = ""

ws.row_dimensions[sig_row].height = 18
ws.row_dimensions[sig_row+1].height = 60

# Bottom border line for stamp
for c in "ABCDEFGH":
    ws[f"{c}{sig_row+1}"].border = Border(bottom=thin)

# ---- Footer ----
foot = sig_row + 3
ws.merge_cells(f"A{foot}:H{foot}")
ws[f"A{foot}"] = ("Bu belge proforma faturadır; ticari fatura niteliği taşımaz. "
                 "Fiyatlar geçerlilik tarihine kadar geçerlidir. "
                 "This is a proforma invoice and does not constitute a commercial invoice.")
ws[f"A{foot}"].font = font_footer
ws[f"A{foot}"].alignment = center
ws.row_dimensions[foot].height = 26

# Print area
ws.print_area = f"A1:H{foot}"

wb.save(OUT)
print(f"Saved: {OUT}")
